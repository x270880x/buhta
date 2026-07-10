#!/usr/bin/env python3
"""Собирает kuvd.xlsx из kuvd.json. Запускать после любой правки kuvd.json:

    python3 tools/make_kuvd_xlsx.py
"""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
k = json.load(open(os.path.join(ROOT, 'kuvd.json'), encoding='utf-8'))

HEAD_FILL = PatternFill('solid', fgColor='1F2430')
HEAD_FONT = Font(bold=True, color='E8C547', size=11)
ST_FILL = {'готово': 'D6F5E3', 'приостановка': 'FDF3D0', 'отказ': 'FBD9D9'}
ST_FONT = {'готово': '1B6B3A', 'приостановка': '8A6D0B', 'отказ': 'A31515'}
THIN = Side(style='thin', color='D9D9D9')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill, cell.font = HEAD_FILL, HEAD_FONT
        cell.alignment = Alignment(vertical='center')
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 22


def autosize(ws, maxw=70):
    for col in ws.columns:
        w = max(len(str(c.value or '')) for c in col) + 2
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w, maxw)


wb = Workbook()

# ── Лист 1: все КУВД ────────────────────────────────────────────────────────
ws = wb.active
ws.title = 'Все КУВД'
ws.append(['№', 'КУВД', 'Владелец', 'Статус', 'Объект', 'Укр. участок',
           'Кадастр РФ', 'Приостановлено до', 'Примечание'])

i = 0
for grp in k['владельцы'] + k['без_владельца']:
    owner = grp['владелец']
    if grp in k['без_владельца']:
        owner = 'НЕ УКАЗАН · ' + owner.replace('без владельца ', '').strip('()')
    for it in grp['номера']:
        i += 1
        notes = []
        if it.get('дубль_в_списке'):
            notes.append('прислан дважды')
        if it.get('⚠'):
            notes.append(it['⚠'])
        ws.append([i, it['кувд'], owner, it.get('статус', ''), it.get('обьект', ''),
                   it.get('укр_участок', ''), it.get('кадастр_рф', ''), it.get('до', ''),
                   '; '.join(notes)])
        st = it.get('статус')
        if st in ST_FILL:
            for c in range(1, 10):
                cell = ws.cell(row=i + 1, column=c)
                cell.fill = PatternFill('solid', fgColor=ST_FILL[st])
            ws.cell(row=i + 1, column=4).font = Font(bold=True, color=ST_FONT[st])
        if notes:
            ws.cell(row=i + 1, column=9).font = Font(color='8A6D0B')

for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=9):
    for cell in row:
        cell.border = BORDER
style_header(ws, 9)
autosize(ws)

# ── Лист 2: сверка по владельцам ────────────────────────────────────────────
ws2 = wb.create_sheet('Сверка')
ws2.append(['Владелец', 'Объектов', 'С зап. кадастром (×1)', 'Без кадастра РФ (×2)',
            'Ожидается КУВД', 'Прислано', 'Расхождение'])
for o in k['владельцы']:
    s = o['сверка']
    d = s['разница']
    ws2.append([o['владелец'], s['обьектов'], s['с_запорожским_кадастром'], s['без_кадастра_рф'],
                s['ожидается_кувд'], s['прислано'], d])
    r = ws2.max_row
    fill = 'D6F5E3' if d == 0 else ('FDF3D0' if d < 0 else 'DCEBFA')
    for c in range(1, 8):
        ws2.cell(row=r, column=c).fill = PatternFill('solid', fgColor=fill)
        ws2.cell(row=r, column=c).border = BORDER
for c in range(1, 8):
    ws2.cell(row=1, column=c).border = BORDER
style_header(ws2, 7)
autosize(ws2)
r = ws2.max_row + 2
ws2.cell(row=r, column=1, value='Правило:').font = Font(bold=True)
ws2.cell(row=r, column=2, value=k['правило_сверки']).alignment = Alignment(wrap_text=True)
ws2.cell(row=r + 1, column=2, value='Расхождение — подсказка, где перепроверить, а не утверждение.')

# ── Лист 3: проблемы ────────────────────────────────────────────────────────
ws3 = wb.create_sheet('Проблемы')
ws3.append(['Тип', 'КУВД', 'Подробности'])
pr = k['проблемы']
for n in pr['битые_номера']:
    ws3.append(['битый номер', n, '6 цифр вместо 7 — вероятна опечатка'])
for x in pr['дубли']:
    ws3.append(['дубль в списке', x['кувд'], f"{x['повторов']} раза · {', '.join(x['владельцы'])}"])
for n in pr['отказ_без_обьекта']:
    ws3.append(['отказ без объекта', n, 'решения нет, к какому участку относится — неизвестно'])
for x in pr['без_владельца']:
    ws3.append(['без владельца', '', x])
for row in ws3.iter_rows(min_row=1, max_row=ws3.max_row, max_col=3):
    for cell in row:
        cell.border = BORDER
style_header(ws3, 3)
autosize(ws3)

out = os.path.join(ROOT, 'kuvd.xlsx')
wb.save(out)
t = k['итого']
print(f"{out}: {t['уникальных']} номеров, 3 листа")
